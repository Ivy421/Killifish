#!/usr/bin/env python
# coding: utf-8

# In[1]:


import numpy as np
import pandas as pd
import os, shutil, re, csv,glob


# # STEP 1. raw data extraction

# In[2]:


def before_during_led():
    import numpy as np
    import pandas as pd
    import os, shutil, re, csv,glob,xlsxwriter, openpyxl
    # target folder name
    sub_folder = ['d48','d77']
    sheetname = ['averaged_before_led_df','dist_vel_acc_df']
    RawExtract = pd.DataFrame()
    # delete previous folder, and create new blank folder to save extraction result
    if os.path.exists(os.getcwd()+'/Raw_Extraction'):
        shutil.rmtree(os.getcwd()+'/Raw_Extraction')
    #os.makedirs(os.getcwd()+'/Raw_Extraction', mode = 0o777, exist_ok=False)
    for subf in sub_folder:
        os.makedirs(os.getcwd()+'/Raw_Extraction/%s'%subf, mode = 0o777, exist_ok=False)
    
    file_pattern = "LED*"
    during_led_df = pd.DataFrame()
    
    # 1. output of avg_before.xlsx
    avg_bf_df = pd.DataFrame()
    for fish_life in sub_folder:
        avg_before_path = os.getcwd()+ "/" + fish_life + "/dabest_plots/paired/averaged_before_led_df.csv"
        avg_bf_df = pd.read_csv(avg_before_path)
        #avg_bf_df.to_csv(os.getcwd()+'/Raw_Extraction/%s/%s_beforeLED.csv'%(fish_life,fish_life))
    
    #for fish_life in sub_folder:
        print('Processing fish group:%s'%fish_life)
        during_sep_file = glob.glob(os.getcwd() + "/"+ fish_life + "/"+ file_pattern)
        loop=0
        for subf in during_sep_file:
            file_list = os.listdir(subf)

            for filename in file_list:
                if filename.endswith('.csv'):
                    #print(filename)
                    loop+=1
                    csv_content = pd.read_csv(subf +"/"+filename)
                    if loop == 1:
                        during_led_df = pd.concat( [during_led_df,csv_content], axis = 1)
                    else:
                        column_idx = csv_content.columns.get_loc("uroa_control")
                        useful_content = csv_content.iloc[:, column_idx+1:]
                        during_led_df = pd.concat( [during_led_df,useful_content], axis = 1)
        # firstly sort by name, secondly sort by day
        during_led_df = during_led_df.sort_values(by= 'name')
        during_led_df = during_led_df.groupby('name').apply(lambda x: x.sort_values(by='day')).reset_index(drop=True)
        #during_led_df.to_csv(os.getcwd()+'/Raw_Extraction/%s/%s_duringLED.csv'%(fish_life,fish_life),index = False)
        with pd.ExcelWriter(os.getcwd()+'/Raw_Extraction/%s/raw_bf_durLED.xlsx'%fish_life, engine='xlsxwriter') as writer:
            avg_bf_df.to_excel(writer, sheet_name='beforeLED', index=False)  # 写入第一个 sheet
            during_led_df.to_excel(writer, sheet_name='duringLED', index=False)
        
        during_led_df = pd.DataFrame()

    #3. output of diff by before to during

    diff_df = pd.DataFrame()
    for fish_life in sub_folder:
        during_led_total = pd.read_excel(os.getcwd()+'/Raw_Extraction/%s/raw_bf_durLED.xlsx'%fish_life, sheet_name = 'duringLED')
        before_led_total = pd.read_excel(os.getcwd()+'/Raw_Extraction/%s/raw_bf_durLED.xlsx'%fish_life, sheet_name = 'beforeLED')
        diff_df = during_led_total.iloc[:,:3]
        dur_col_name =  during_led_total.columns[3:].tolist() #list
        for colname in dur_col_name:
            if colname in before_led_total.columns:
                diff_col_data = during_led_total[colname] - before_led_total[colname]
                diff_df = pd.concat([diff_df,diff_col_data],axis = 1)
        with pd.ExcelWriter(os.getcwd()+'/Raw_Extraction/%s/raw_bf_durLED.xlsx'%fish_life, engine='openpyxl',mode = 'a') as writer:
            diff_df.to_excel(writer, sheet_name='diffLED', index=False)   
        dur_col_name = []
        diff_df = pd.DataFrame()
    return 'process done'


# # STEP 2 calculation of z-score

# In[3]:


def z_score_cal():
    import numpy as np
    import pandas as pd
    import os, shutil, re, csv,glob
    sub_folder = ['d48','d77']
    file_pattern = "raw_bf*"
    diff_led = pd.DataFrame()
    z_score = pd.DataFrame()
    diff_file_path = []
    for subf in sub_folder:
        diff_file_path = glob.glob(os.getcwd() + "/Raw_Extraction/%s/"%subf + file_pattern)
        print('Z_score calculation from:%s'%diff_file_path)
        avg_list = []
        std_list = []  
        diff_file = pd.read_excel(diff_file_path[0],sheet_name = 'diffLED')
        z_score = diff_file.iloc[:,:3]
        colname = diff_file.columns
        #print(colname)
        for j in range (diff_file.shape[1] - 3):
            data = diff_file.iloc[:, j+3]
            avg_list.append(data.mean())
            std_list.append(data.std())
            z_score['z_%s'%colname[j+3]] = (diff_file.iloc[:,j+3]-data.mean())/data.std()
        z_score = pd.concat([z_score,pd.DataFrame(avg_list,columns = ['avg']),pd.DataFrame(std_list,columns = ['std'])], axis = 1)
        z_score.to_csv(os.getcwd()+'/Raw_Extraction/%s/%s_z_score.csv'%(subf,subf),index = False)
    return 'xx_z_score.xlsx  FILE DONE!'


# # STEP 3 calculation of z diff

# In[5]:


def diff_zscore_byday():
    import numpy as np
    import pandas as pd
    import os, shutil, re, csv,glob
    sub_folder = ['d48','d77']
    file_pattern = "*z_score*"
    df_move_col = pd.DataFrame()
    for subf in sub_folder:
        zscore_file_path = glob.glob(os.getcwd() + "/Raw_Extraction/%s/"%subf + file_pattern)
        merge_group_z = pd.DataFrame()
        z_score_ori = pd.read_csv(zscore_file_path[0])
        z_score_ori = z_score_ori.iloc[:,:-2]
        name_grouped = z_score_ori.groupby('name')
        for name, group in name_grouped:
            if group.shape[0] == 4:
                diffD1D2 = group.iloc[0,3:]-group.iloc[1,3:]
                diffD1D3 = group.iloc[0,3:]-group.iloc[2,3:]
                diffD2D3 = group.iloc[1,3:]-group.iloc[2,3:]
                diffD1D4 = group.iloc[0,3:]-group.iloc[3,3:]
                diffD2D4 = group.iloc[1,3:]-group.iloc[3,3:]

            # 3 days data,maybe day123，day134，day234, calculate by case
            elif group.shape[0] == 3:

                if 1 not in group['day'].tolist(): #D2D3,D2D4
                    #print("D2D3,D2D4:\n",group['day'])
                    diffD2D3 = group.iloc[0,3:]-group.iloc[1,3:]
                    diffD2D4 = group.iloc[0,3:]-group.iloc[2,3:]
                    diffD1D2 = diffD1D3 = diffD1D4 = group.iloc[0,3:]-group.iloc[0,3:]
                elif 2 not in group['day'].tolist(): #D1D3 D1D4
                    #print("D1D3 D1D4:\n",group['day'])
                    diffD1D3 = group.iloc[0,3:]-group.iloc[1,3:]
                    diffD1D4 = group.iloc[0,3:]-group.iloc[2,3:]
                    diffD1D2 = diffD2D3 = diffD2D4 = group.iloc[0,3:]-group.iloc[0,3:]
                elif 3 not in group['day'].tolist(): #D1D2,D1D4,D2D4
                    #print("D1D2,D1D4,D2D4:\n",group['day'])
                    diffD1D2 = group.iloc[0,3:]-group.iloc[1,3:]
                    diffD1D4 = group.iloc[0,3:]-group.iloc[2,3:]
                    diffD2D4 = group.iloc[1,3:]-group.iloc[2,3:]
                    diffD1D3 = diffD2D3 = group.iloc[0,3:]-group.iloc[0,3:]
                elif 4 not in group['day'].tolist():#D1D2,D1D3,D2D3
                    #print("D1D2,D1D3,D2D3:\n",group['day'])
                    diffD1D2 = group.iloc[0,3:]-group.iloc[1,3:]
                    diffD1D3 = group.iloc[0,3:]-group.iloc[2,3:]
                    diffD2D3 = group.iloc[1,3:]-group.iloc[2,3:]
                    diffD2D4 = diffD1D4 = group.iloc[0,3:]-group.iloc[0,3:]

            # only 2 days data，classify the calculation: 12，13，14，23，24，34
            elif group.shape[0] == 2: 
                if 1 and 2 in group['day'].tolist():
                    #print('day1 and 2:\n',group['day'])
                    diffD2D4 = diffD1D4 = diffD2D3 = diffD1D3 = group.iloc[0,3:]-group.iloc[0,3:]
                    diffD1D2 = group.iloc[0,3:]-group.iloc[1,3:]
                elif 1 and 3 in group['day'].tolist():
                    #print('day1 and 3:\n',group['day'])
                    diffD2D4 = diffD1D4 = diffD2D3 = diffD1D2 = group.iloc[0,3:]-group.iloc[0,3:]
                    diffD1D3 = group.iloc[0,3:]-group.iloc[1,3:]
                elif 1 and 4 in group['day'].tolist():
                    #print('day1 and 3:\n',group['day'])
                    diffD2D4 = diffD1D3 = diffD2D3 = diffD1D2 = group.iloc[0,3:]-group.iloc[0,3:]
                    diffD1D4 = group.iloc[0,3:]-group.iloc[1,3:]

                elif 2 and 3 in group['day'].tolist():
                    #print('day1 and 3:\n',group['day'])
                    diffD2D4 = diffD1D4 = diffD1D3 = diffD1D2 = group.iloc[0,3:]-group.iloc[0,3:]
                    diffD2D3 = group.iloc[0,3:]-group.iloc[1,3:]            
                elif 2 and 4 in group['day'].tolist():
                    #print('day1 and 3:\n',group['day'])
                    diffD1D3 = diffD1D4 = diffD2D3 = diffD1D2 = group.iloc[0,3:]-group.iloc[0,3:]
                    diffD2D4 = group.iloc[0,3:]-group.iloc[1,3:]            
                else:
                    diffD2D4 = diffD1D4 = diffD2D3 = diffD1D3 = diffD1D2 = group.iloc[0,3:]-group.iloc[0,3:]
            else:
                diffD2D4 = diffD1D4 = diffD2D3 = diffD1D3 = diffD1D2 = group.iloc[0,3:]-group.iloc[0,3:]

            one_group_diffZ = pd.concat([diffD1D2, diffD1D3,diffD1D4, diffD2D3, diffD2D4], axis=0)
            one_group_diffZ = pd.DataFrame(one_group_diffZ)
            one_group_diffZ = one_group_diffZ.transpose()
            one_group_diffZ['name'] = name
            one_group_diffZ = one_group_diffZ.set_index('name') 
            merge_group_z = pd.concat([merge_group_z,one_group_diffZ],axis = 0)

        merge_group_z = merge_group_z.abs()

        # rename columns by adding D1D2 etc.
        new_columns = []
        for i, col in enumerate(merge_group_z.columns):
            if i < 8:
                new_columns.append("D1D2" + col)
            elif i>=8 and i< 16:
                new_columns.append("D1D3" + col)
            elif i>=16 and i< 24:
                new_columns.append("D1D4" + col)
            elif i>=24 and i< 32:
                new_columns.append("D2D3" + col)
            elif i>=32:
                new_columns.append("D2D4" + col)
        merge_group_z.columns = new_columns

        # remove columns to get a new df
        df_move_col = pd.DataFrame(index = merge_group_z.index)
        for i in range(int(merge_group_z.shape[1]/5)):
            k=0
            for j in range(int(merge_group_z.shape[1]/8)):
            # 0,8,16,24,32
                df_move_col = pd.concat([df_move_col,merge_group_z.iloc[:,k+i]],axis = 1)
                k+=8

        df_move_col.to_excel(os.getcwd()+'/Raw_Extraction/%s/%s_diffZ.xlsx'%(subf,subf),index = True)

    return 'xx_diffZ.xlsx FILE DONE!'


# In[7]:


def Top_diffZ():
    import numpy as np
    import pandas as pd
    import os, shutil, re, csv,glob
    import openpyxl
    from openpyxl.styles import Font
    sub_folder = ['d48','d77']
    for fish_life in sub_folder:
        diff_z_ori = pd.read_excel(os.getcwd()+'/Raw_Extraction/%s/%s_diffZ.xlsx'%(fish_life,fish_life))
        wb = openpyxl.load_workbook(os.getcwd()+'/Raw_Extraction/%s/%s_diffZ.xlsx'%(fish_life,fish_life))
        sheet = wb['Sheet1']
        fish_top_33 = round(diff_z_ori.shape[0]*0.33)
        col_idx = 1
        for col in diff_z_ori.columns[1:]:
            topvalues = diff_z_ori.nlargest(fish_top_33, col)
            target_val = topvalues.iloc[:,col_idx].values
            for row in sheet.iter_rows(min_col=col_idx+1, max_col=col_idx+1):
                for cell in row:
                    if cell.value in target_val:
                        cell.font = Font(color='FF0000')
            col_idx += 1
        wb.save(os.getcwd()+'/Raw_Extraction/%s/%s_diffZ.xlsx'%(fish_life,fish_life) )     
    return ' red for top 33% data'


# ## STEP4  count how many red font cell in each row

# In[9]:


def count_red():
    # count red text for each row:
    import numpy as np
    import pandas as pd
    import os, shutil, re, csv,glob
    import openpyxl
    from openpyxl.styles import Font
    sub_folder = ['d48','d77']
    for fish_life in sub_folder:
        red_count_list = []
        diff_z_ori = pd.read_excel(os.getcwd()+'/Raw_Extraction/%s/%s_diffZ.xlsx'%(fish_life,fish_life))
        wb = openpyxl.load_workbook(os.getcwd()+'/Raw_Extraction/%s/%s_diffZ.xlsx'%(fish_life,fish_life))
        sheet = wb['Sheet1']
        for i in range(2,diff_z_ori.shape[0]):
            row_cells = list(sheet[i])
            red_count = 0
            for cell in row_cells:
                if cell.font and cell.font.color and cell.font.color.rgb == '00FF0000':
                    red_count += 1
            print(i,red_count)
            red_count_list.append(red_count)
        count_col = diff_z_ori.shape[1]+1
        for i in range(len(red_count_list)):
            if i == 0:
                sheet.cell(row=1, column=count_col, value='Count_Red')
            sheet.cell(row=i+2, column=count_col, value=red_count_list[i])
        wb.save(os.getcwd()+'/Raw_Extraction/%s/%s_diffZ.xlsx'%(fish_life,fish_life)) 
    
    return "diffZ.xlsx FILE DONE!!"


# In[ ]:




